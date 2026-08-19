import streamlit as st
import cv2
import time
import pandas as pd
import winsound
import wave
import math
import struct
import threading

from pathlib import Path
import sys
from datetime import datetime

from ultralytics import YOLO
from openpyxl import Workbook, load_workbook


# ============================================================
# PROJECT PATH
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]

sys.path.append(str(PROJECT_ROOT))

from utils.violation_logger import (
    log_violation,
    get_log_file
)


# ============================================================
# DIRECTORIES
# ============================================================

ALARM_DIR = PROJECT_ROOT / "alarm"
RECORDS_DIR = PROJECT_ROOT / "records"
SNAPSHOT_DIR = RECORDS_DIR / "snapshots"

ALARM_DIR.mkdir(exist_ok=True)
RECORDS_DIR.mkdir(exist_ok=True)
SNAPSHOT_DIR.mkdir(exist_ok=True)


# ============================================================
# FILES
# ============================================================

MODEL_PATH = PROJECT_ROOT / "models" / "best.pt"

SIREN_FILE = ALARM_DIR / "generated_siren.wav"

EXCEL_FILE = RECORDS_DIR / "ppe_violation_records.xlsx"


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="Industrial PPE Command Center",
    page_icon="🦺",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ============================================================
# PROFESSIONAL CSS
# ============================================================

st.markdown(
    """
    <style>

    .main {
        background-color: #0e1117;
    }

    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
    }

    .dashboard-title {
        font-size: 2.2rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }

    .dashboard-subtitle {
        color: #9ca3af;
        font-size: 1rem;
        margin-bottom: 1.5rem;
    }

    .kpi-card {
        background: #171b24;
        border: 1px solid #2a3040;
        border-radius: 12px;
        padding: 18px;
        min-height: 120px;
    }

    .kpi-title {
        color: #9ca3af;
        font-size: 0.85rem;
        margin-bottom: 8px;
    }

    .kpi-value {
        font-size: 2rem;
        font-weight: 700;
    }

    .kpi-green {
        color: #22c55e;
    }

    .kpi-red {
        color: #ef4444;
    }

    .kpi-yellow {
        color: #f59e0b;
    }

    .kpi-blue {
        color: #38bdf8;
    }

    .status-running {
        background: rgba(34,197,94,0.12);
        border: 1px solid rgba(34,197,94,0.35);
        border-radius: 10px;
        padding: 12px 16px;
        color: #22c55e;
        font-weight: 600;
    }

    .status-stopped {
        background: rgba(156,163,175,0.12);
        border: 1px solid rgba(156,163,175,0.25);
        border-radius: 10px;
        padding: 12px 16px;
        color: #d1d5db;
        font-weight: 600;
    }

    .alarm-active {
        background: rgba(239,68,68,0.15);
        border: 1px solid rgba(239,68,68,0.5);
        border-radius: 10px;
        padding: 14px;
        color: #ef4444;
        font-weight: 700;
        text-align: center;
    }

    .alarm-inactive {
        background: rgba(34,197,94,0.12);
        border: 1px solid rgba(34,197,94,0.35);
        border-radius: 10px;
        padding: 14px;
        color: #22c55e;
        font-weight: 700;
        text-align: center;
    }

    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# SESSION STATE
# ============================================================

if "monitoring" not in st.session_state:
    st.session_state.monitoring = False

if "camera" not in st.session_state:
    st.session_state.camera = None

if "alarm_active" not in st.session_state:
    st.session_state.alarm_active = False

if "last_logged" not in st.session_state:
    st.session_state.last_logged = {}

if "total_session_violations" not in st.session_state:
    st.session_state.total_session_violations = 0


# ============================================================
# PPE CONFIGURATION
# ============================================================

REQUIRED_PPE = {
    "helmet": "Helmet",
    "gloves": "Gloves",
    "vest": "Safety Vest",
    "boots": "Safety Boots",
    "goggles": "Safety Goggles"
}


VIOLATION_CLASSES = {
    "no_helmet": "Helmet",
    "no_gloves": "Gloves",
    "no_boots": "Safety Boots",
    "no_goggle": "Safety Goggles"
}


# ============================================================
# MODEL
# ============================================================

@st.cache_resource
def load_model():

    return YOLO(
        str(MODEL_PATH)
    )


model = load_model()


# ============================================================
# CREATE SIREN
# ============================================================

def create_siren():

    if SIREN_FILE.exists():
        return

    sample_rate = 44100
    duration = 3.0
    volume = 0.95

    with wave.open(
        str(SIREN_FILE),
        "w"
    ) as wav:

        wav.setnchannels(1)
        wav.setsampwidth(2)
        wav.setframerate(sample_rate)

        total_samples = int(
            sample_rate * duration
        )

        for i in range(total_samples):

            t = i / sample_rate

            cycle = t % 1.5

            if cycle < 0.75:

                progress = cycle / 0.75

                frequency = (
                    650 +
                    (1350 - 650) *
                    progress
                )

            else:

                progress = (
                    cycle - 0.75
                ) / 0.75

                frequency = (
                    1350 -
                    (1350 - 650) *
                    progress
                )

            sample = math.sin(
                2 *
                math.pi *
                frequency *
                t
            )

            sample += (
                0.30 *
                math.sin(
                    2 *
                    math.pi *
                    frequency *
                    2 *
                    t
                )
            )

            sample += (
                0.12 *
                math.sin(
                    2 *
                    math.pi *
                    frequency *
                    3 *
                    t
                )
            )

            sample *= volume

            sample = max(
                -1,
                min(
                    1,
                    sample
                )
            )

            wav.writeframes(
                struct.pack(
                    "<h",
                    int(
                        sample *
                        32767
                    )
                )
            )


create_siren()


# ============================================================
# ALARM
# ============================================================

alarm_thread = None
alarm_stop_event = threading.Event()


def start_alarm():

    global alarm_thread

    if st.session_state.alarm_active:
        return

    st.session_state.alarm_active = True

    alarm_stop_event.clear()

    def play():

        while not alarm_stop_event.is_set():

            try:

                winsound.PlaySound(
                    str(SIREN_FILE),
                    winsound.SND_FILENAME
                )

            except Exception:

                break

    alarm_thread = threading.Thread(
        target=play,
        daemon=True
    )

    alarm_thread.start()


def stop_alarm():

    global alarm_thread

    alarm_stop_event.set()

    st.session_state.alarm_active = False

    try:

        winsound.PlaySound(
            None,
            winsound.SND_PURGE
        )

    except Exception:

        pass

    alarm_thread = None


# ============================================================
# EXCEL
# ============================================================

def initialize_excel():

    if EXCEL_FILE.exists():
        return

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "PPE Violations"

    worksheet.append([
        "Timestamp",
        "Person ID",
        "Missing PPE",
        "Violation Count",
        "Status",
        "Snapshot"
    ])

    workbook.save(
        EXCEL_FILE
    )


initialize_excel()


# ============================================================
# SAVE SNAPSHOT
# ============================================================

def save_snapshot(
    frame,
    person_id
):

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    filename = (
        f"person_{person_id}_"
        f"{timestamp}.jpg"
    )

    path = SNAPSHOT_DIR / filename

    cv2.imwrite(
        str(path),
        frame
    )

    return path


# ============================================================
# EXCEL LOG
# ============================================================

def write_violation_to_excel(
    person_id,
    missing_ppe,
    frame
):

    initialize_excel()

    try:

        snapshot_path = save_snapshot(
            frame,
            person_id
        )

        workbook = load_workbook(
            EXCEL_FILE
        )

        worksheet = workbook[
            "PPE Violations"
        ]

        timestamp = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        missing_text = ", ".join(
            sorted(missing_ppe)
        )

        worksheet.append([
            timestamp,
            f"Person-{person_id}",
            missing_text,
            len(missing_ppe),
            "NON-COMPLIANT",
            str(snapshot_path)
        ])

        workbook.save(
            EXCEL_FILE
        )

        st.session_state.total_session_violations += 1

    except PermissionError:

        pass

    except Exception:

        pass


# ============================================================
# IOU
# ============================================================

def calculate_iou(
    box_a,
    box_b
):

    ax1, ay1, ax2, ay2 = box_a

    bx1, by1, bx2, by2 = box_b

    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)

    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)

    iw = max(
        0,
        ix2 - ix1
    )

    ih = max(
        0,
        iy2 - iy1
    )

    intersection = iw * ih

    area_a = (
        max(
            0,
            ax2 - ax1
        )
        *
        max(
            0,
            ay2 - ay1
        )
    )

    area_b = (
        max(
            0,
            bx2 - bx1
        )
        *
        max(
            0,
            by2 - by1
        )
    )

    union = (
        area_a +
        area_b -
        intersection
    )

    if union <= 0:
        return 0

    return intersection / union


# ============================================================
# BOX CENTER
# ============================================================

def box_center(box):

    x1, y1, x2, y2 = box

    return (
        (x1 + x2) / 2,
        (y1 + y2) / 2
    )


# ============================================================
# POINT INSIDE
# ============================================================

def point_inside_box(
    point,
    box
):

    x, y = point

    x1, y1, x2, y2 = box

    return (
        x1 <= x <= x2
        and
        y1 <= y <= y2
    )


# ============================================================
# ASSIGN PPE
# ============================================================

def assign_ppe_to_person(
    ppe_box,
    persons
):

    center = box_center(
        ppe_box
    )

    best_person = None
    best_score = 0

    for person in persons:

        person_box = person["box"]

        if point_inside_box(
            center,
            person_box
        ):

            score = 1.0

        else:

            score = calculate_iou(
                ppe_box,
                person_box
            )

        if score > best_score:

            best_score = score

            best_person = person

    return best_person


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:

    st.markdown(
        "## 🦺 Control Center"
    )

    st.divider()

    confidence = st.slider(
        "Detection Confidence",
        0.10,
        0.90,
        0.40,
        0.05
    )

    st.divider()

    st.markdown(
        "### System Status"
    )

    if st.session_state.monitoring:

        st.success(
            "🟢 Monitoring ACTIVE"
        )

    else:

        st.info(
            "⚪ Monitoring STOPPED"
        )

    if st.session_state.alarm_active:

        st.error(
            "🔴 ALARM ACTIVE"
        )

    else:

        st.success(
            "🟢 Alarm Ready"
        )

    st.divider()

    st.caption(
        "AI Engine: YOLO"
    )

    st.caption(
        "Processing: CPU"
    )

    st.caption(
        "Tracking: ByteTrack"
    )


# ============================================================
# HEADER
# ============================================================

st.markdown(
    '<div class="dashboard-title">'
    '🦺 Industrial PPE Command Center'
    '</div>',
    unsafe_allow_html=True
)

st.markdown(
    '<div class="dashboard-subtitle">'
    'Real-time workplace safety monitoring and '
    'PPE compliance management'
    '</div>',
    unsafe_allow_html=True
)


# ============================================================
# LOAD EXISTING DATA
# ============================================================

if EXCEL_FILE.exists():

    try:

        dashboard_df = pd.read_excel(
            EXCEL_FILE
        )

    except Exception:

        dashboard_df = pd.DataFrame()

else:

    dashboard_df = pd.DataFrame()


# ============================================================
# KPI CALCULATIONS
# ============================================================

if not dashboard_df.empty:

    total_violations = len(
        dashboard_df
    )

    workers_involved = (
        dashboard_df[
            "Person ID"
        ].nunique()
    )

else:

    total_violations = 0
    workers_involved = 0


# ============================================================
# KPI CARDS
# ============================================================

k1, k2, k3, k4 = st.columns(4)


with k1:

    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-title">
                👥 WORKERS INVOLVED
            </div>
            <div class="kpi-value kpi-blue">
                {workers_involved}
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


with k2:

    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-title">
                🚨 TOTAL VIOLATIONS
            </div>
            <div class="kpi-value kpi-red">
                {total_violations}
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


with k3:

    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-title">
                ⚠️ SESSION VIOLATIONS
            </div>
            <div class="kpi-value kpi-yellow">
                {st.session_state.total_session_violations}
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


with k4:

    alarm_text = (
        "ACTIVE"
        if st.session_state.alarm_active
        else "READY"
    )

    alarm_class = (
        "kpi-red"
        if st.session_state.alarm_active
        else "kpi-green"
    )

    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-title">
                🔊 ALARM
            </div>
            <div class="kpi-value {alarm_class}">
                {alarm_text}
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


st.write("")


# ============================================================
# CONTROLS
# ============================================================

control1, control2, control3 = st.columns(
    [1, 1, 2]
)


with control1:

    if st.button(
        "▶️ START MONITORING",
        type="primary",
        use_container_width=True
    ):

        st.session_state.monitoring = True


with control2:

    if st.button(
        "⛔ STOP MONITORING",
        use_container_width=True
    ):

        st.session_state.monitoring = False

        stop_alarm()

        if (
            st.session_state.camera
            is not None
        ):

            st.session_state.camera.release()

            st.session_state.camera = None


with control3:

    if st.session_state.monitoring:

        st.markdown(
            '<div class="status-running">'
            '🟢 LIVE MONITORING ACTIVE'
            '</div>',
            unsafe_allow_html=True
        )

    else:

        st.markdown(
            '<div class="status-stopped">'
            '⚪ MONITORING STOPPED'
            '</div>',
            unsafe_allow_html=True
        )


st.divider()


# ============================================================
# LIVE MONITORING SECTION
# ============================================================

live_col, info_col = st.columns(
    [2.3, 1]
)


with live_col:

    st.subheader(
        "📹 Live Safety Camera"
    )

    frame_placeholder = st.empty()


with info_col:

    st.subheader(
        "🚨 Safety Status"
    )

    status_placeholder = st.empty()

    alarm_placeholder = st.empty()

    alert_placeholder = st.empty()

    stats_placeholder = st.empty()


# ============================================================
# MONITORING
# ============================================================

if st.session_state.monitoring:

    if st.session_state.camera is None:

        camera = cv2.VideoCapture(0)

        if not camera.isOpened():

            st.error(
                "Camera could not be opened."
            )

            st.session_state.monitoring = False

            st.stop()

        st.session_state.camera = camera


    camera = st.session_state.camera


    try:

        while st.session_state.monitoring:

            success, frame = camera.read()

            if not success:

                break


            # =================================================
            # YOLO
            # =================================================

            results = model.track(

                source=frame,

                conf=confidence,

                persist=True,

                tracker="bytetrack.yaml",

                device="cpu",

                verbose=False

            )


            result = results[0]

            persons = []

            ppe_detections = []


            # =================================================
            # READ BOXES
            # =================================================

            if result.boxes is not None:

                boxes = result.boxes

                for index in range(
                    len(boxes)
                ):

                    class_id = int(
                        boxes.cls[index]
                    )

                    class_name = model.names[
                        class_id
                    ]

                    coords = (

                        boxes.xyxy[index]

                        .cpu()

                        .numpy()

                    )

                    x1, y1, x2, y2 = map(
                        int,
                        coords
                    )

                    detection_box = [
                        x1,
                        y1,
                        x2,
                        y2
                    ]


                    # =========================================
                    # PERSON
                    # =========================================

                    if class_name == "Person":

                        if boxes.id is not None:

                            track_id = int(
                                boxes.id[index]
                            )

                        else:

                            track_id = 0


                        persons.append({

                            "id": track_id,

                            "box": detection_box,

                            "ppe": set(),

                            "violations": set()

                        })


                    # =========================================
                    # PPE
                    # =========================================

                    elif (
                        class_name in REQUIRED_PPE
                        or
                        class_name in VIOLATION_CLASSES
                    ):

                        ppe_detections.append({

                            "class": class_name,

                            "box": detection_box

                        })


            # =================================================
            # ASSIGN PPE
            # =================================================

            for detection in ppe_detections:

                person = assign_ppe_to_person(

                    detection["box"],

                    persons

                )

                if person is None:
                    continue


                class_name = detection[
                    "class"
                ]


                if class_name in REQUIRED_PPE:

                    person["ppe"].add(
                        class_name
                    )

                elif class_name in VIOLATION_CLASSES:

                    person[
                        "violations"
                    ].add(
                        VIOLATION_CLASSES[
                            class_name
                        ]
                    )


            # =================================================
            # STATISTICS
            # =================================================

            compliant_people = 0

            violating_people = 0

            total_missing = 0

            violation_messages = []


            # =================================================
            # WORKERS
            # =================================================

            for person in persons:

                person_id = person["id"]

                detected_ppe = person["ppe"]

                explicit_violations = (
                    person["violations"]
                )


                missing_ppe = set(

                    REQUIRED_PPE[item]

                    for item in REQUIRED_PPE

                    if item not in detected_ppe

                )


                missing_ppe.update(
                    explicit_violations
                )


                is_violation = (
                    len(missing_ppe) > 0
                )


                x1, y1, x2, y2 = (
                    person["box"]
                )


                # =============================================
                # VIOLATION
                # =============================================

                if is_violation:

                    violating_people += 1

                    total_missing += (
                        len(missing_ppe)
                    )


                    color = (
                        0,
                        0,
                        255
                    )


                    status_text = (
                        f"ID {person_id}: "
                        "VIOLATION"
                    )


                    missing_text = ", ".join(

                        sorted(
                            missing_ppe
                        )

                    )


                    violation_messages.append(

                        f"🚨 Worker #{person_id}\n"
                        f"Missing: {missing_text}"

                    )


                    # =========================================
                    # LOG COOLDOWN
                    # =========================================

                    current_time = time.time()


                    log_key = (

                        person_id,

                        tuple(
                            sorted(
                                missing_ppe
                            )
                        )

                    )


                    previous_time = (
                        st.session_state
                        .last_logged
                        .get(
                            log_key,
                            0
                        )
                    )


                    if (
                        current_time -
                        previous_time
                        >= 5
                    ):

                        try:

                            log_violation(

                                f"Person-{person_id}",

                                sorted(
                                    missing_ppe
                                )

                            )

                        except Exception:

                            pass


                        write_violation_to_excel(

                            person_id,

                            missing_ppe,

                            frame

                        )


                        st.session_state.last_logged[
                            log_key
                        ] = current_time


                else:

                    compliant_people += 1

                    color = (
                        0,
                        255,
                        0
                    )

                    status_text = (
                        f"ID {person_id}: "
                        "COMPLIANT"
                    )


                # =============================================
                # DRAW WORKER
                # =============================================

                cv2.rectangle(

                    frame,

                    (x1, y1),

                    (x2, y2),

                    color,

                    3

                )


                cv2.putText(

                    frame,

                    status_text,

                    (
                        x1,
                        max(
                            y1 - 10,
                            20
                        )
                    ),

                    cv2.FONT_HERSHEY_SIMPLEX,

                    0.65,

                    color,

                    2

                )


                # =============================================
                # MISSING PPE
                # =============================================

                if is_violation:

                    text_y = y1 + 25


                    for missing in sorted(
                        missing_ppe
                    ):

                        cv2.putText(

                            frame,

                            f"Missing: {missing}",

                            (
                                x1,
                                text_y
                            ),

                            cv2.FONT_HERSHEY_SIMPLEX,

                            0.5,

                            color,

                            2

                        )

                        text_y += 22


            # =================================================
            # ALARM / STATUS
            # =================================================

            if violating_people > 0:

                start_alarm()


                status_placeholder.markdown(

                    '<div class="alarm-active">'
                    '🔴 PPE VIOLATION DETECTED'
                    '</div>',

                    unsafe_allow_html=True

                )


                alarm_placeholder.markdown(

                    '<div class="alarm-active">'
                    '🔊 EMERGENCY ALARM ACTIVE'
                    '</div>',

                    unsafe_allow_html=True

                )


                if violation_messages:

                    alert_placeholder.warning(

                        "\n\n".join(
                            violation_messages
                        )

                    )


            else:

                stop_alarm()


                status_placeholder.markdown(

                    '<div class="alarm-inactive">'
                    '🟢 ALL WORKERS COMPLIANT'
                    '</div>',

                    unsafe_allow_html=True

                )


                alarm_placeholder.markdown(

                    '<div class="alarm-inactive">'
                    '🔇 ALARM OFF'
                    '</div>',

                    unsafe_allow_html=True

                )


                alert_placeholder.empty()


            # =================================================
            # LIVE STATS
            # =================================================

            with stats_placeholder.container():

                st.metric(
                    "👥 Workers",
                    len(persons)
                )

                st.metric(
                    "🟢 Compliant",
                    compliant_people
                )

                st.metric(
                    "🔴 Violating",
                    violating_people
                )

                st.metric(
                    "⚠️ Missing PPE",
                    total_missing
                )


            # =================================================
            # DISPLAY
            # =================================================

            frame_rgb = cv2.cvtColor(
                frame,
                cv2.COLOR_BGR2RGB
            )


            frame_placeholder.image(

                frame_rgb,

                channels="RGB",

                use_container_width=True

            )


            time.sleep(0.03)


    finally:

        camera.release()

        st.session_state.camera = None

        stop_alarm()


else:

    frame_placeholder.info(
        "📷 Camera is currently offline. "
        "Click START MONITORING to begin."
    )


# ============================================================
# MANAGER ANALYTICS
# ============================================================

st.divider()

st.header(
    "📊 Manager Analytics"
)


if EXCEL_FILE.exists():

    try:

        df = pd.read_excel(
            EXCEL_FILE
        )

    except Exception:

        df = pd.DataFrame()

else:

    df = pd.DataFrame()


if not df.empty:

    analytics1, analytics2 = st.columns(2)


    # ========================================================
    # VIOLATIONS BY PPE
    # ========================================================

    with analytics1:

        st.subheader(
            "⚠️ Violations by PPE Type"
        )


        ppe_counts = {}


        for item in REQUIRED_PPE.values():

            ppe_counts[item] = 0


        for value in df[
            "Missing PPE"
        ].astype(str):

            for item in ppe_counts:

                if item in value:

                    ppe_counts[item] += 1


        chart_df = pd.DataFrame({

            "PPE": list(
                ppe_counts.keys()
            ),

            "Violations": list(
                ppe_counts.values()
            )

        })


        st.bar_chart(

            chart_df.set_index(
                "PPE"
            )

        )


    # ========================================================
    # WORKER VIOLATIONS
    # ========================================================

    with analytics2:

        st.subheader(
            "👤 Violations by Worker"
        )


        worker_counts = (
            df[
                "Person ID"
            ]
            .value_counts()
            .head(10)
        )


        st.bar_chart(
            worker_counts
        )


else:

    st.info(
        "Analytics will appear after "
        "the first PPE violation."
    )


# ============================================================
# VIOLATION HISTORY
# ============================================================

st.divider()

st.header(
    "📋 Violation History"
)


if not df.empty:

    st.dataframe(

        df.iloc[::-1],

        use_container_width=True,

        height=400

    )


    # ========================================================
    # DOWNLOAD EXCEL
    # ========================================================

    with open(
        EXCEL_FILE,
        "rb"
    ) as file:

        st.download_button(

            label="⬇️ Download Excel Report",

            data=file,

            file_name=(
                "industrial_ppe_violation_report.xlsx"
            ),

            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )

        )


else:

    st.info(
        "No violation records available."
    )


# ============================================================
# SNAPSHOT INFORMATION
# ============================================================

st.divider()

st.header(
    "📸 Violation Evidence"
)


snapshot_files = list(
    SNAPSHOT_DIR.glob("*.jpg")
)


if snapshot_files:

    st.write(
        f"{len(snapshot_files)} violation "
        "snapshot(s) stored."
    )


    selected_snapshot = st.selectbox(

        "Select violation snapshot",

        snapshot_files,

        format_func=lambda x: x.name

    )


    if selected_snapshot:

        st.image(

            str(selected_snapshot),

            caption=selected_snapshot.name,

            use_container_width=True

        )


else:

    st.info(
        "Violation snapshots will appear here "
        "after violations are detected."
    )